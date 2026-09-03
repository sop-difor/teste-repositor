#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_stg_sinapi.py — Lê a planilha mensal do SINAPI (Caixa, .xlsx) e gera o CSV
no formato da tabela de recebimento `stg_sinapi`.

Abas usadas:  ISD (insumo onerado), ICD (insumo desonerado),
              CSD (composição onerada), CCD (composição desonerada),
              Analítico (composição x itens x coeficientes).
Ignora: Menu, Busca, ISE, CSE, "Analítico com Custo".

Regras:
  - Preço = coluna do estado CE; se CE for 0/vazio, usa a coluna SP.
  - O código das composições (CSD/CCD) vem de uma fórmula HYPERLINK(...,NNNN) —
    o script extrai o número da fórmula.
  - `composicao` (linhas 'C') = JSON dos itens da aba Analítico:
    [{"codigo_item","descricao_item","tipo_item","coeficiente","unidade"}, ...]

Saída (colunas da stg_sinapi):
  identificacao, codigo, descricao, unidade, preco_unitario, tipo_encargo,
  referencia, composicao, origem_preco
  (origem_preco = 'CE' normalmente, 'SP' quando o preço de Ceará estava zerado
   e caímos no de São Paulo, '' quando não há preço em nenhum dos dois.)

Uso:
  python gerar_stg_sinapi.py "CAMINHO\\SINAPI_Referência_2025_01.xlsx" 2025-01
  ->  gera  stg_sinapi_2025-01.csv  nesta pasta

Requisitos:  pip install openpyxl
"""

import csv
import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta a biblioteca openpyxl. Rode:  pip install openpyxl")


# --------------------------------------------------------------------------- #
def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).strip().upper()


_num_re = re.compile(r"^-?\d+(?:[.,]\d+)?$")


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not _num_re.match(s):
        return None
    return float(s.replace(".", "").replace(",", ".")) if "," in s else float(s)


_hl_re = re.compile(r"HYPERLINK\s*\(.*?,\s*\"?(\d{2,})\"?\s*\)\s*$", re.I | re.S)


def extrai_codigo(valor_cache, valor_formula):
    """código pode ser: número direto, ou fórmula =HYPERLINK("#"...,NNNN)."""
    for v in (valor_formula, valor_cache):
        if v is None:
            continue
        s = str(v).strip()
        if s.startswith("="):
            m = _hl_re.search(s)
            if m:
                return m.group(1)
            m2 = re.search(r"(\d{3,})", s)   # último recurso
            if m2:
                return m2.group(1)
        elif re.match(r"^\d+(\.0)?$", s):
            return s.split(".")[0]
    return None


def achar_aba(wb, prefixo):
    p = norm(prefixo)
    for nome in wb.sheetnames:
        if norm(nome) == p or norm(nome).startswith(p):
            return nome
    return None


def achar_header(linhas, obrig, limite=25):
    """devolve (i, {NOME_NORM: col}) — mapa junta a linha achada + 8 acima + 3 abaixo
    (o cabeçalho do SINAPI espalha siglas de estado em linhas separadas)."""
    for i in range(min(limite, len(linhas))):
        m0 = {}
        for j, c in enumerate(linhas[i]):
            nc = norm(c)
            if nc:
                m0.setdefault(nc, j)
        if all(any(o in k for k in m0) for o in obrig):
            mapa = {}
            for r in linhas[max(0, i - 8): i + 4]:
                for j, c in enumerate(r):
                    nc = norm(c)
                    if nc:
                        mapa.setdefault(nc, j)
            return i, mapa
    return None, None


def col(mapa, *nomes, exato=False):
    for n in nomes:
        a = norm(n)
        if a in mapa:
            return mapa[a]
    if not exato:
        for n in nomes:
            a = norm(n)
            for k, j in mapa.items():
                if a and a in k:
                    return j
    return None


# --------------------------------------------------------------------------- #
def ler_itens(rows_v, rows_f, titulo):
    """ISD/ICD/CSD/CCD -> {codigo: {'descricao','unidade','preco'}} com regra CE->SP."""
    i, mapa = achar_header(rows_v, ["CODIGO", "DESCRICAO", "UNIDADE"])
    if i is None:
        raise RuntimeError(f"[{titulo}] cabeçalho não encontrado.")
    c_cod = col(mapa, "CODIGO DO INSUMO", "CODIGO DA COMPOSICAO", "CODIGO")
    c_des = col(mapa, "DESCRICAO DO INSUMO", "DESCRICAO DA COMPOSICAO", "DESCRICAO")
    c_uni = col(mapa, "UNIDADE")
    c_ce = col(mapa, "CE", exato=True)
    c_sp = col(mapa, "SP", exato=True)
    if None in (c_cod, c_des, c_uni, c_ce):
        raise RuntimeError(f"[{titulo}] colunas faltando: "
                           f"cod={c_cod} des={c_des} uni={c_uni} CE={c_ce}. "
                           f"headers={sorted(mapa)[:60]}")
    out = {}
    n_sem_preco = 0
    for rv, rf in zip(rows_v[i + 1:], rows_f[i + 1:]):
        cod = extrai_codigo(rv[c_cod] if c_cod < len(rv) else None,
                            rf[c_cod] if c_cod < len(rf) else None)
        if not cod:
            continue
        preco = to_num(rv[c_ce]) if c_ce < len(rv) else None
        origem = "CE"
        if (preco is None or preco == 0) and c_sp is not None and c_sp < len(rv):
            sp = to_num(rv[c_sp])
            if sp is not None and sp != 0:
                preco, origem = sp, "SP"
        if preco is None or preco == 0:
            # sem preço em CE nem em SP -> descarta a linha (não vai para o banco)
            n_sem_preco += 1
            continue
        des = rv[c_des] if c_des < len(rv) else None
        uni = rv[c_uni] if c_uni < len(rv) else None
        out[cod] = {
            "descricao": str(des).strip() if des is not None else "",
            "unidade": str(uni).strip() if uni is not None else "",
            "preco": preco,
            "origem": origem,
        }
    n_sp = sum(1 for v in out.values() if v["origem"] == "SP")
    print(f"  {titulo}: {len(out)} itens  (CE col {c_ce}, SP col {c_sp}) — "
          f"{n_sp} via SP, {n_sem_preco} descartados sem preço")
    return out


def ler_analitico(rows_v, rows_f):
    i, mapa = achar_header(rows_v, ["COMPOSICAO", "ITEM", "COEFICIENTE"])
    if i is None:
        raise RuntimeError("[Analítico] cabeçalho não encontrado.")
    c_comp = col(mapa, "CODIGO DA COMPOSICAO", "CODIGO COMPOSICAO")
    c_tipo = col(mapa, "TIPO ITEM", "TIPO DO ITEM")
    c_icod = col(mapa, "CODIGO DO ITEM", "CODIGO ITEM")
    c_ides = col(mapa, "DESCRICAO DO ITEM", "DESCRICAO ITEM", "DESCRICAO")
    c_iuni = col(mapa, "UNIDADE DO ITEM", "UNIDADE ITEM", "UNIDADE")
    c_coef = col(mapa, "COEFICIENTE")
    if None in (c_comp, c_icod, c_coef):
        raise RuntimeError(f"[Analítico] colunas faltando: comp={c_comp} item={c_icod} "
                           f"coef={c_coef}. headers={sorted(mapa)[:60]}")
    out = {}
    comp = None
    for rv, rf in zip(rows_v[i + 1:], rows_f[i + 1:]):
        cc = extrai_codigo(rv[c_comp] if c_comp < len(rv) else None,
                           rf[c_comp] if c_comp < len(rf) else None)
        if cc:
            comp = cc
        icod = extrai_codigo(rv[c_icod] if c_icod < len(rv) else None,
                             rf[c_icod] if c_icod < len(rf) else None)
        if comp is None or not icod:
            continue
        tp = norm(rv[c_tipo]) if (c_tipo is not None and c_tipo < len(rv)) else ""
        out.setdefault(comp, []).append({
            "tipo_item": "COMPOSICAO" if tp.startswith("COMP") else "INSUMO",
            "codigo_item": icod,
            "descricao_item": (str(rv[c_ides]).strip()
                               if c_ides is not None and c_ides < len(rv) and rv[c_ides] is not None else ""),
            "unidade": (str(rv[c_iuni]).strip()
                        if c_iuni is not None and c_iuni < len(rv) and rv[c_iuni] is not None else ""),
            "coeficiente": to_num(rv[c_coef]) if c_coef < len(rv) else None,
        })
    tot = sum(len(v) for v in out.values())
    print(f"  Analítico: {len(out)} composições, {tot} linhas de item")
    return out


# --------------------------------------------------------------------------- #
def main():
    if len(sys.argv) < 2:
        sys.exit('Uso: python gerar_stg_sinapi.py "<arquivo.xlsx>" [AAAA-MM]')
    caminho = Path(sys.argv[1])
    if not caminho.exists():
        sys.exit(f"Arquivo não encontrado: {caminho}")

    mes = sys.argv[2].strip() if len(sys.argv) >= 3 and sys.argv[2].strip() else ""
    if not mes:                       # tenta descobrir pelo nome do arquivo
        m = re.search(r"(20\d{2})[ _.-]?(0[1-9]|1[0-2])(?!\d)", caminho.stem)
        if m:
            mes = f"{m.group(1)}-{m.group(2)}"
            print(f"Mês detectado pelo nome do arquivo: {mes}")
    if not re.match(r"^\d{4}-\d{2}$", mes):
        sys.exit("Não consegui descobrir o mês pelo nome do arquivo. "
                 "Passe como 2º argumento no formato AAAA-MM (ex.: 2025-01).")
    referencia = f"{mes}-01"

    print(f"Lendo {caminho.name} (pode levar ~30 s)...")
    wb_v = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    wb_f = openpyxl.load_workbook(caminho, read_only=True, data_only=False)
    print("Abas:", wb_v.sheetnames)

    def rows(wb, prefixo):
        nome = achar_aba(wb, prefixo)
        if nome is None:
            sys.exit(f"Aba '{prefixo}' não encontrada.")
        return [r for r in wb[nome].iter_rows(values_only=True)]

    dados = {}
    for pref, tipo_enc in (("ISD", "onerada"), ("ICD", "desonerada"),
                           ("CSD", "onerada"), ("CCD", "desonerada")):
        rv, rf = rows(wb_v, pref), rows(wb_f, pref)
        dados[pref] = ler_itens(rv, rf, pref)
    ana = ler_analitico(rows(wb_v, "Analítico"), rows(wb_f, "Analítico"))

    isd, icd, csd, ccd = dados["ISD"], dados["ICD"], dados["CSD"], dados["CCD"]

    def item_filho(cod, tipo_item, onerado):
        fonte = (isd if onerado else icd) if tipo_item == "INSUMO" else (csd if onerado else ccd)
        return fonte.get(cod)

    def monta_comp(cod, onerado):
        itens = ana.get(cod)
        if not itens:
            return ""
        arr = []
        for it in itens:
            f = item_filho(it["codigo_item"], it["tipo_item"], onerado)
            arr.append({
                "codigo_item": it["codigo_item"],
                "descricao_item": it["descricao_item"],
                "tipo_item": it["tipo_item"],
                "coeficiente": it["coeficiente"],
                "unidade": it["unidade"],
                "preco_unitario": f["preco"] if f else None,
                "origem_preco": f["origem"] if f else "",
            })
        return json.dumps(arr, ensure_ascii=False)

    linhas = []
    for cod, it in isd.items():
        linhas.append(["I", cod, it["descricao"], it["unidade"], it["preco"], "onerada", referencia, "", it["origem"]])
    for cod, it in icd.items():
        linhas.append(["I", cod, it["descricao"], it["unidade"], it["preco"], "desonerada", referencia, "", it["origem"]])
    for cod, it in csd.items():
        linhas.append(["C", cod, it["descricao"], it["unidade"], it["preco"], "onerada", referencia,
                       monta_comp(cod, True), it["origem"]])
    for cod, it in ccd.items():
        linhas.append(["C", cod, it["descricao"], it["unidade"], it["preco"], "desonerada", referencia,
                       monta_comp(cod, False), it["origem"]])

    saida = Path(__file__).with_name(f"stg_sinapi_{mes}.csv")
    with open(saida, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["identificacao", "codigo", "descricao", "unidade",
                    "preco_unitario", "tipo_encargo", "referencia", "composicao", "origem_preco"])
        for r in linhas:
            w.writerow(["" if c is None else c for c in r])

    c_com = sum(1 for r in linhas if r[0] == "C" and r[7])
    c_sem = sum(1 for r in linhas if r[0] == "C" and not r[7])
    print(f"\nOK -> {saida}")
    print(f"  {len(linhas)} linhas | insumos {len(isd)+len(icd)} | composições {len(csd)+len(ccd)}"
          f" (com analítico {c_com}, sem {c_sem})")
    print(f"  linhas sem preço em CE nem SP foram descartadas (ver contagem por aba acima).")
    print(f"\nDepois:  TRUNCATE stg_sinapi;  importe o CSV;  SELECT rt_aplicar_sinapi('{referencia}');")


if __name__ == "__main__":
    main()
